import argparse
import json
import re
from pathlib import Path

import openpyxl


PROVIDER_NORMALIZATION = {
    'S-DHL': 'Singapore-DHL',
    'S-UPS': 'Singapore-UPS',
    'D-DHL': 'DUBAI-DHL',
    'D-UPS': 'DUBAI-UPS',
    'S-FedEx': 'Singapore-FedEx',
    'D-FedEx': 'DUBAI-FedEx',
    'Master': 'Master',
    'MASTER': 'Master',
    'Master-SF': 'Master-SF',
    'Master -SF': 'Master-SF',
    'Master- SF': 'Master-SF',
    'Master-Nice': 'Master-Nice',
    'D-DHL/Risk': 'DUBAI-DHL/Risk',
    'OCS': 'OCS',
}


def read_float(value):
    if value is None:
        return None

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_text(value):
    if value is None:
        return ''
    return str(value).strip()


def extract_country_code(country_text):
    text = read_text(country_text)
    match = re.search(r'\(([A-Z]{2})\)?', text)
    if match:
        return match.group(1)
    return None


def extract_header_weight(header_text):
    text = read_text(header_text).lower()
    if not text or 'kg' not in text:
        return None
    if 'per' in text:
        return None

    match = re.search(r'(\d+(?:\.\d+)?)', text)
    if not match:
        return None

    return read_float(match.group(1))


def extract_zone_number(zone_text):
    text = read_text(zone_text)
    if not text:
        return None

    match = re.search(r'(\d+)', text)
    if not match:
        return None

    return int(match.group(1))


def detect_rate_zone_columns(ws, header_row):
    zone_columns = []

    for col_idx in range(1, ws.max_column + 1):
        header_text = read_text(ws.cell(header_row, col_idx).value)
        if not header_text:
            continue

        normalized = header_text.lower().replace(' ', '')
        if normalized.startswith('zone'):
            zone_number = extract_zone_number(header_text)
            if zone_number is not None:
                zone_columns.append((zone_number, col_idx))

    return zone_columns


def parse_dhl_zone_list(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]

    country_headers = []
    zone_headers = []
    header_row = 2

    best_header_row = None
    best_country_count = -1
    scan_header_end = min(ws.max_row, 12)

    for candidate_row in range(1, scan_header_end + 1):
        candidate_country_count = 0
        candidate_zone_count = 0

        for col_idx in range(1, ws.max_column + 1):
            header_text = read_text(ws.cell(candidate_row, col_idx).value).lower()
            if not header_text:
                continue

            if 'countries' in header_text:
                candidate_country_count += 1
            elif header_text.startswith('zone'):
                candidate_zone_count += 1

        if candidate_country_count > 0 and candidate_zone_count > 0 and candidate_country_count > best_country_count:
            best_country_count = candidate_country_count
            best_header_row = candidate_row

    if best_header_row is not None:
        header_row = best_header_row

    for col_idx in range(1, ws.max_column + 1):
        header_text = read_text(ws.cell(header_row, col_idx).value)
        if not header_text:
            continue

        normalized = header_text.lower()
        if 'countries' in normalized:
            country_headers.append(col_idx)
        elif normalized.startswith('zone'):
            zone_headers.append(col_idx)

    header_pairs = []
    def pick_zone_value_col(zone_header_col):
        candidates = [zone_header_col]
        if zone_header_col + 1 <= ws.max_column:
            candidates.append(zone_header_col + 1)
        if zone_header_col - 1 >= 1:
            candidates.append(zone_header_col - 1)

        best_col = zone_header_col
        best_score = -1
        scan_end = min(ws.max_row, 80)

        for candidate_col in candidates:
            numeric_count = 0
            for row_idx in range(5, scan_end + 1):
                if read_float(ws.cell(row_idx, candidate_col).value) is not None:
                    numeric_count += 1

            if numeric_count > best_score:
                best_score = numeric_count
                best_col = candidate_col

        return best_col

    for index, country_col in enumerate(country_headers):
        next_country_col = country_headers[index + 1] if index + 1 < len(country_headers) else ws.max_column + 1
        candidate_zones = [col for col in zone_headers if country_col < col < next_country_col]
        if not candidate_zones:
            candidate_zones = [col for col in zone_headers if col > country_col]
        zone_col = candidate_zones[0] if candidate_zones else None
        if zone_col is not None:
            header_pairs.append((country_col, pick_zone_value_col(zone_col)))

    zones = []
    for row_idx in range(3, ws.max_row + 1):
        for country_col, zone_col in header_pairs:
            country_text = read_text(ws.cell(row_idx, country_col).value)
            zone_value = read_float(ws.cell(row_idx, zone_col).value)

            if not country_text or zone_value is None:
                continue

            country_code = extract_country_code(country_text)
            if country_code is None:
                continue

            country_name = re.sub(r'\s*\(.*$', '', country_text).strip()
            zones.append({
                'country_code': country_code,
                'country_name': country_name,
                'zone': int(zone_value),
            })

    return {
        'zones': zones,
        'total_zones': len(zones),
    }


def parse_dhl_rates(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]

    doc_header_row = None
    non_doc_header_row = None
    per_kg_header_row = None

    for row_idx in range(1, min(ws.max_row + 1, 100)):
        val = str(ws.cell(row_idx, 1).value or '').strip().lower()
        if val == 'kg' or 'weight' in val:
            if doc_header_row is None:
                doc_header_row = row_idx
            elif non_doc_header_row is None:
                non_doc_header_row = row_idx
        elif 'multiplier' in val or 'per kg' in val or 'per\nkg' in val or (val and 'kg' in val and '-' in val and doc_header_row):
            if per_kg_header_row is None and 'kg' not in val: # Avoid matching the KG header
                per_kg_header_row = row_idx
            elif 'multiplier' in val or 'per kg' in val:
                per_kg_header_row = row_idx

    # If per_kg is not found by text, try to find the row after the last non_document slab
    if doc_header_row is None: doc_header_row = 4
    if non_doc_header_row is None: non_doc_header_row = 10

    document_zone_columns = detect_rate_zone_columns(ws, doc_header_row)
    non_document_zone_columns = detect_rate_zone_columns(ws, non_doc_header_row)

    if not non_document_zone_columns:
        non_document_zone_columns = document_zone_columns

    document_rows = []
    non_document_rows = []
    per_kg_bands = []

    for row_idx in range(doc_header_row + 1, min(doc_header_row + 20, ws.max_row + 1)):
        val = str(ws.cell(row_idx, 1).value or '').strip().lower()
        if 'non-document' in val or 'non document' in val or val == 'kg':
            break
        weight = read_float(ws.cell(row_idx, 1).value)
        if weight is None:
            continue
        zones = {}
        for zone_number, col_idx in document_zone_columns:
            zones[zone_number] = read_float(ws.cell(row_idx, col_idx).value)
        document_rows.append({'weight': weight, 'zones': zones})

    last_non_doc_row = non_doc_header_row
    for row_idx in range(non_doc_header_row + 1, min(non_doc_header_row + 100, ws.max_row + 1)):
        val = str(ws.cell(row_idx, 1).value or '').strip().lower()
        if 'multiplier' in val or 'per kg' in val or ('-' in val and 'kg' not in val and not val.replace('.', '').isdigit()):
            break
        weight = read_float(ws.cell(row_idx, 1).value)
        if weight is None:
            if val: # Hit a text row (like 'Multiplier')
                break
            continue
        zones = {}
        for zone_number, col_idx in non_document_zone_columns:
            zones[zone_number] = read_float(ws.cell(row_idx, col_idx).value)
        non_document_rows.append({'weight': weight, 'zones': zones})
        last_non_doc_row = row_idx

    if per_kg_header_row is None:
        # Fallback for per_kg: scan rows after the last non-document slab
        for row_idx in range(last_non_doc_row + 1, min(last_non_doc_row + 15, ws.max_row + 1)):
            from_w = read_float(ws.cell(row_idx, 1).value)
            if from_w is not None and from_w >= 30:
                per_kg_header_row = row_idx
                break

    if per_kg_header_row:
        for row_idx in range(per_kg_header_row, min(per_kg_header_row + 10, ws.max_row + 1)):
            from_w = read_float(ws.cell(row_idx, 1).value)
            to_w = read_float(ws.cell(row_idx, 2).value)
            
            if from_w is not None:
                zones = {}
                for zone_number, col_idx in non_document_zone_columns:
                    zones[zone_number] = read_float(ws.cell(row_idx, col_idx).value)
                
                per_kg_bands.append({
                    'from': from_w,
                    'to': to_w,
                    'zones': zones,
                })

    return {
        'document': document_rows,
        'non_document': non_document_rows,
        'per_kg_bands': per_kg_bands,
    }


def parse_aramex_zones(path):
    if not path:
        return {'zones': [], 'total_zones': 0}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]
    zones = []
    for row in ws.iter_rows(values_only=True):
        for i in range(0, len(row), 2):
            if i + 1 < len(row):
                country = row[i]
                zone = row[i+1]
                if isinstance(country, str) and isinstance(zone, str) and len(zone.strip()) == 1:
                    country_clean = re.sub(r'[^a-zA-Z\s]', '', country).strip()
                    if country_clean and country_clean.lower() != 'country':
                        zones.append({
                            "country_code": "", # will be matched later or ignored
                            "country_name": country_clean,
                            "zone": ord(zone.strip().upper()) - 64
                        })
    return {'zones': zones, 'total_zones': len(zones)}


def extract_aramex_zone_positions(header_row):
    positions = []

    for index, cell in enumerate(header_row[1:], start=1):
        header = read_text(cell).upper()
        if re.fullmatch(r'[A-N]', header):
            positions.append(index)

    return positions

def parse_aramex_rates_up_to_10(path):
    if not path:
        return {'document': [], 'non_document': []}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]
    document_rows = []
    non_document_rows = []
    mode = 'document'
    zone_positions = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        weight_col = str(row[0]).strip().lower()
        if 'non-document' in weight_col:
            mode = 'non_document'
            zone_positions = []
            continue
        if weight_col in ('kg', 'cost') or not weight_col:
            if weight_col == 'kg':
                zone_positions = extract_aramex_zone_positions(row)
            continue
            
        weight_str = str(row[0]).strip()
        weight_val = None
        if weight_str == 'ADD 0.5':
            weight_val = 'ADD 0.5' # Special token for document increment
        else:
            weight_val = read_float(row[0])
            if weight_val is None:
                continue
                
        slab_zones = {}
        active_positions = zone_positions or list(range(1, min(15, len(row))))
        for zone_number, position in enumerate(active_positions, start=1):
            if position < len(row):
                val = read_float(row[position])
                if val is not None:
                    slab_zones[zone_number] = val
                    
        if slab_zones:
            if mode == 'document':
                document_rows.append({'weight': weight_val, 'zones': slab_zones})
            else:
                non_document_rows.append({'weight': weight_val, 'zones': slab_zones})
                
    return {'document': document_rows, 'non_document': non_document_rows}

def parse_aramex_rates_above_10(path):
    if not path:
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]
    data = []
    
    headers = []
    header_positions = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        first_cell = read_text(row[0]).lower()
        if first_cell.startswith('per kg'):
            headers = [read_text(cell) for cell in row]
            header_positions = [index for index, cell in enumerate(row) if index > 0 and read_text(cell)]
            break
            
    if not headers:
        return []

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        band = str(row[0]).strip().lower()
        if not band.endswith('kg+'):
            continue
        match = re.search(r'\d+', band)
        if not match:
            continue
        weight_band = float(match.group())
            
        rates = {}
        active_positions = header_positions or list(range(1, min(len(row), len(headers))))
        for position in active_positions:
            if position >= len(row) or position >= len(headers):
                continue

            price = read_float(row[position])
            if price is not None:
                header = headers[position]
                countries = [c.strip() for c in header.replace('\n', '/').split('/') if c.strip()]
                for c in countries:
                    rates[c] = price
        if rates:
            data.append({"weight_band": weight_band, "rates": rates})
    return data

def parse_ups(path):
    if not path:
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    
    ws = None
    if 'online' in wb.sheetnames:
        ws = wb['online']
    else:
        for sheet_name in wb.sheetnames:
            candidate = wb[sheet_name]
            for row in candidate.iter_rows(min_row=1, max_row=20, values_only=True):
                if row and row[0] and 'country' in str(row[0]).strip().lower():
                    ws = candidate
                    break
            if ws:
                break
        if not ws:
            ws = wb[wb.sheetnames[0]]
            
    header_row = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0]:
            first_cell = str(row[0]).strip().lower()
            if 'country' in first_cell or 'destination' in first_cell:
                header_row = row
                break
            
    if not header_row:
        return []

    countries_per_col = {}
    for col_idx, cell in enumerate(header_row):
        if col_idx == 0 or not cell:
            continue
        text = str(cell).strip()
        countries = [c.strip() for c in text.split('/') if c.strip()]
        countries_per_col[col_idx] = countries

    def parse_weight_band(label, previous_range_end=None):
        text = str(label).strip().lower()

        if 'kg' not in text:
            return None

        range_match = re.search(r'(\d+(?:\.\d+)?)\s*[-–]\s*(\d+(?:\.\d+)?)', text)
        if range_match:
            start_weight = float(range_match.group(1))
            end_weight = float(range_match.group(2))
            return start_weight, end_weight

        plus_match = re.search(r'(\d+(?:\.\d+)?)\s*\+', text)
        if plus_match:
            start_weight = float(plus_match.group(1))

            if previous_range_end is not None and abs(previous_range_end - start_weight) < 0.001:
                return start_weight + 0.1, start_weight

            return start_weight, start_weight

        match = re.search(r'^(\d+(?:\.\d+)?)', text)
        if not match:
            return None

        weight_band = float(match.group(1))
        return weight_band, weight_band

    rates = []
    previous_range_end = None
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue

        parsed_band = parse_weight_band(row[0], previous_range_end)
        if parsed_band is None:
            continue

        weight_band, range_end = parsed_band
        band_rates = {}
        for col_idx, countries in countries_per_col.items():
            if col_idx < len(row):
                price = read_float(row[col_idx])
                if price is not None:
                    for c in countries:
                        band_rates[c] = price

        if band_rates:
            rates.append({'weight_band': weight_band, 'rates': band_rates})

            previous_range_end = range_end

    return rates


def parse_tge(path):
    if not path:
        return []

    wb = openpyxl.load_workbook(path, data_only=True)

    ws = wb['ONLINE'] if 'ONLINE' in wb.sheetnames else wb[wb.sheetnames[0]]

    rates = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue

        band_text = str(row[0]).strip().lower()
        if 'kg' not in band_text:
            continue

        match = re.search(r'^(\d+(?:\.\d+)?)', band_text)
        if not match:
            continue

        weight_band = float(match.group(1))
        rate = read_float(row[1] if len(row) > 1 else None)
        if rate is None:
            continue

        rates.append({
            'weight_band': weight_band,
            'rates': {
                'Australia': rate,
            },
        })

    return rates


def extract_weight_slab(weight_text):
    text = read_text(weight_text).lower()
    if not text:
        return None

    match = re.search(r'(\d+(?:\.\d+)?)', text)
    if not match:
        return None

    return read_float(match.group(1))


def parse_ocs(path):
    if not path:
        return {
            'country_code': 'JP',
            'country_name': 'Japan',
            'rates': [],
        }

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    country_name = 'Japan'
    rates = []
    add_0_5 = None
    per_21_kg = None
    per_31_kg = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first_cell = read_text(row[0] if row else None)
        second_cell = read_float(row[1] if row and len(row) > 1 else None)

        if not first_cell:
            continue

        normalized = first_cell.lower()

        if row_idx <= 3 and normalized not in {'ocs', 'weight', 'amount'} and second_cell is None:
            country_name = first_cell.strip().title()
            continue

        if normalized in {'weight', 'amount'}:
            continue

        if 'next - 0.50kg' in normalized or 'next-0.50kg' in normalized:
            add_0_5 = second_cell
            continue

        if 'flat rate' in normalized and '+21kg' in normalized:
            per_21_kg = second_cell
            continue

        if 'flat rate' in normalized and '+31kg' in normalized:
            per_31_kg = second_cell
            continue

        weight = extract_weight_slab(first_cell)
        if weight is None or second_cell is None:
            continue

        if weight <= 10.0:
            rates.append({
                'weight': weight,
                'price': second_cell,
            })

    return {
        'country_code': 'JP',
        'country_name': country_name,
        'rates': rates,
        'add_0_5': add_0_5,
        'per_21_kg': per_21_kg,
        'per_31_kg': per_31_kg,
    }



def parse_master(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 1'] if 'Table 1' in wb.sheetnames else wb[wb.sheetnames[0]]

    data = {
        'providers': {},
        'total_rates': 0,
    }

    header_row = 7
    expected_parcel_weights = [0.5 + (i * 0.5) for i in range(20)]

    parcel_col_by_weight = {}
    per_0_5_col = None
    per_21_col = None
    per_31_col = None

    for col_idx in range(6, ws.max_column + 1):
        header_text = read_text(ws.cell(header_row, col_idx).value)
        if not header_text:
            continue

        lower = header_text.lower()

        if 'per' in lower and '0.5' in lower:
            per_0_5_col = col_idx
            continue

        if '21' in lower and 'kg+' in lower:
            per_21_col = col_idx
            continue

        if '31' in lower and 'kg+' in lower:
            per_31_col = col_idx
            continue

        weight = extract_header_weight(header_text)
        if weight is None:
            continue

        normalized_weight = round(weight, 1)
        if normalized_weight in expected_parcel_weights and normalized_weight not in parcel_col_by_weight:
            parcel_col_by_weight[normalized_weight] = col_idx

    current_country = None
    for row_idx in range(9, ws.max_row + 1):
        destination = ws.cell(row_idx, 1).value
        provider_raw = ws.cell(row_idx, 2).value

        if destination and str(destination).strip():
            current_country = str(destination).strip()

        if provider_raw is None or current_country is None:
            continue

        provider_text = str(provider_raw).strip()
        if provider_text in ('Service', 'Provider'):
            continue

        provider_full = PROVIDER_NORMALIZATION.get(provider_text)
        if provider_full is None:
            continue

        doc_0_5 = read_float(ws.cell(row_idx, 3).value)
        doc_1_0 = read_float(ws.cell(row_idx, 4).value)
        doc_add_0_5 = read_float(ws.cell(row_idx, 5).value)

        parcel_rates = {}
        for i in range(20):
            weight_slab = 0.5 + (i * 0.5)
            col_idx = parcel_col_by_weight.get(round(weight_slab, 1), 6 + i)
            parcel_rates[str(weight_slab)] = read_float(ws.cell(row_idx, col_idx).value)

        per_0_5_kg = read_float(ws.cell(row_idx, per_0_5_col or 26).value)
        per_21_kg = read_float(ws.cell(row_idx, per_21_col or 27).value)
        per_31_kg = read_float(ws.cell(row_idx, per_31_col or 28).value)

        if provider_full not in data['providers']:
            data['providers'][provider_full] = {'countries': {}}

        data['providers'][provider_full]['countries'][current_country] = {
            'document': {
                '0.5': doc_0_5,
                '1.0': doc_1_0,
                'add_0.5': doc_add_0_5,
            },
            'parcel': parcel_rates,
            'per_0_5_kg': per_0_5_kg,
            'per_21_kg': per_21_kg,
            'per_31_kg': per_31_kg,
        }
        data['total_rates'] += 1

    return data


def parse_fedex_zones(path):
    """Parse FedEx Bangladesh zone list (Country -> Zone letter A-Q = 1-17)."""
    if not path:
        return {'zones': [], 'total_zones': 0}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    zones = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue

        country_text = read_text(row[0]).strip()
        zone_col_b = read_text(row[1]).strip().upper() if len(row) > 1 and row[1] else ''

        # Skip section headers (single letter rows like 'A', 'B', …) and blanks
        if not country_text:
            continue
        if len(country_text) == 1 and country_text.isalpha():
            continue
        if not zone_col_b:
            continue

        # Some cells have the zone letter embedded at the end of the country name
        # e.g. "Taiwan F", "Tanzania J", "Turks & Caicos I I"
        # Strip trailing " X" where X is a single uppercase letter
        zone_letter = zone_col_b
        cleaned_country = re.sub(r'\s+[A-Q]$', '', country_text).strip()

        # If the embedded letter matches col B zone we used the cleaned name;
        # otherwise just use the raw country text (no embedded zone suffix)
        embedded = re.search(r'\s+([A-Q])$', country_text)
        if embedded and embedded.group(1) == zone_col_b:
            country_text = cleaned_country

        # Convert zone letter to integer (A=1, B=2, …, Q=17)
        if len(zone_letter) != 1 or not zone_letter.isalpha():
            continue

        zone_number = ord(zone_letter) - ord('A') + 1

        zones.append({
            'country_name': country_text,
            'zone_letter': zone_letter,
            'zone': zone_number,
        })

    return {'zones': zones, 'total_zones': len(zones)}


def parse_fedex_rates(path):
    """Parse FedEx Bangladesh rate sheet.

    Structure:
      Row 3 : header  – 'Weight kg', 'A', 'B', … 'Q'
      Row 4 : 'Documents Rate (FedEx Envelope)' label
      Row 5 : 0.5 kg  document rate per zone
      Row 6 : blank
      Rows 7-11  : document rates for 0.5-2.5 kg (extended doc slabs)
      Row 12: 'Non-Document' label
      Rows 13-31: non-document slab rates  0.5-9.5 kg
      Row 32: '10kg+' label
      Rows 33-39: per-kg band rates (range labels like '10-19', '20-29', …)
    """
    if not path:
        return {'document': [], 'non_document': [], 'per_kg_bands': []}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))

    # --- Detect zone columns from header row (row index 2, 1-based row 3) ---
    header = rows[2]  # 0-indexed row 2 = row 3
    zone_cols = {}  # zone_number -> 0-based column index
    for col_i, cell in enumerate(header):
        text = read_text(cell).strip().upper()
        if len(text) == 1 and text.isalpha() and text >= 'A':
            zone_number = ord(text) - ord('A') + 1
            zone_cols[zone_number] = col_i

    def row_to_zone_rates(row):
        """Extract {zone_number: price} from a data row using detected columns."""
        zones = {}
        for zone_num, col_i in zone_cols.items():
            if col_i < len(row):
                price = read_float(row[col_i])
                if price is not None:
                    zones[zone_num] = price
        return zones

    document_rows = []
    non_document_rows = []
    per_kg_bands = []

    mode = None  # 'document', 'non_document', 'per_kg'
    previous_weight = None

    for row_idx, row in enumerate(rows):
        if not row or row[0] is None:
            continue

        first = read_text(row[0]).strip().lower()

        # Section labels
        if 'documents rate' in first or 'document rate' in first:
            mode = 'document'
            previous_weight = None
            continue
        if first == 'non-document' or first == 'non document':
            mode = 'non_document'
            previous_weight = None
            continue
        if '10kg+' in first or '10 kg+' in first:
            mode = 'per_kg'
            previous_weight = None
            continue

        if mode == 'per_kg':
            # Rows like '10-19', '20-29', '30-39', '40-49', '50-70', '71-100', '101kg+'
            match = re.search(r'(\d+)', first)
            if not match:
                continue
            from_weight = float(match.group(1))
            zones = row_to_zone_rates(row)
            if zones:
                per_kg_bands.append({'from_weight': from_weight, 'zones': zones})
            continue

        # Numeric weight slab row
        weight = read_float(row[0])
        if weight is None:
            continue

        if mode == 'document' and previous_weight is not None and weight < previous_weight:
            mode = 'non_document'
            
        previous_weight = weight

        zones = row_to_zone_rates(row)
        if not zones:
            continue

        if mode == 'document':
            document_rows.append({'weight': weight, 'zones': zones})
        elif mode == 'non_document':
            non_document_rows.append({'weight': weight, 'zones': zones})

    return {
        'document': document_rows,
        'non_document': non_document_rows,
        'per_kg_bands': per_kg_bands,
    }


def main():
    parser = argparse.ArgumentParser(description='Parse DHL + Master monthly rate sheets to JSON')
    parser.add_argument('--dhl', help='Path to DHL rate Excel file')
    parser.add_argument('--dhl-rate', help='Path to DHL rate Excel file')
    parser.add_argument('--dhl-zone', help='Path to DHL zone Excel file')
    parser.add_argument('--master', required=False, help='Path to Master Excel file')
    parser.add_argument('--output', required=True, help='Output JSON path')
    parser.add_argument('--aramex-zone', help='Path to Aramex zone file')
    parser.add_argument('--aramex-upto10', help='Path to Aramex rates up to 10kg file')
    parser.add_argument('--aramex-above10', help='Path to Aramex rates above 10kg file')
    parser.add_argument('--ups', help='Path to UPS rate file')
    parser.add_argument('--ocs', help='Path to OCS Japan rate file')
    parser.add_argument('--tge', help='Path to TGE (Australia) rate file')
    parser.add_argument('--fedex-zone', help='Path to FedEx Bangladesh zone file')
    parser.add_argument('--fedex-rates', help='Path to FedEx Bangladesh rate file')
    args = parser.parse_args()

    dhl_rate_path = Path(args.dhl_rate or args.dhl) if (args.dhl_rate or args.dhl) else None
    dhl_zone_path = Path(args.dhl_zone) if args.dhl_zone else None
    master_path = Path(args.master) if args.master else None
    output_path = Path(args.output)
    aramex_zone_path = Path(args.aramex_zone) if args.aramex_zone else None
    aramex_upto10_path = Path(args.aramex_upto10) if args.aramex_upto10 else None
    aramex_above10_path = Path(args.aramex_above10) if args.aramex_above10 else None
    ups_path = Path(args.ups) if args.ups else None
    ocs_path = Path(args.ocs) if args.ocs else None
    tge_path = Path(args.tge) if args.tge else None
    fedex_zone_path = Path(args.fedex_zone) if args.fedex_zone else None
    fedex_rates_path = Path(args.fedex_rates) if args.fedex_rates else None

    if dhl_rate_path is not None and not dhl_rate_path.exists():
        raise FileNotFoundError(f'DHL rate file not found: {dhl_rate_path}')

    if dhl_zone_path is not None and not dhl_zone_path.exists():
        raise FileNotFoundError(f'DHL zone file not found: {dhl_zone_path}')

    if master_path is not None and not master_path.exists():
        raise FileNotFoundError(f'Master file not found: {master_path}')

    parsed = {
        'dhl': parse_dhl_rates(str(dhl_rate_path)) if dhl_rate_path else {'document': [], 'non_document': [], 'per_kg_bands': []},
        'dhl_zones': parse_dhl_zone_list(str(dhl_zone_path)) if dhl_zone_path is not None else {'zones': [], 'total_zones': 0},
        'master': parse_master(str(master_path)) if master_path else {'providers': {}, 'total_rates': 0},
        'aramex_zones': parse_aramex_zones(str(aramex_zone_path)) if aramex_zone_path else {'zones': [], 'total_zones': 0},
        'aramex_upto10': parse_aramex_rates_up_to_10(str(aramex_upto10_path)) if aramex_upto10_path else {'document': [], 'non_document': []},
        'aramex_above10': parse_aramex_rates_above_10(str(aramex_above10_path)) if aramex_above10_path else [],
        'ups': parse_ups(str(ups_path)) if ups_path else [],
        'ocs': parse_ocs(str(ocs_path)) if ocs_path else {'country_code': 'JP', 'country_name': 'Japan', 'rates': []},
        'tge': parse_tge(str(tge_path)) if tge_path else [],
        'fedex_zones': parse_fedex_zones(str(fedex_zone_path)) if fedex_zone_path else {'zones': [], 'total_zones': 0},
        'fedex': parse_fedex_rates(str(fedex_rates_path)) if fedex_rates_path else {'document': [], 'non_document': [], 'per_kg_bands': []},
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(parsed, indent=2), encoding='utf-8')

    print('Parsing complete')
    print(f"DHL document slabs: {len(parsed['dhl']['document'])}")
    print(f"DHL non-document slabs: {len(parsed['dhl']['non_document'])}")
    print(f"DHL zone mappings: {parsed['dhl_zones']['total_zones']}")
    print(f"Master entries: {parsed['master']['total_rates']}")
    print(f"FedEx zone mappings: {parsed['fedex_zones']['total_zones']}")
    print(f"FedEx document slabs: {len(parsed['fedex']['document'])}")
    print(f"FedEx non-document slabs: {len(parsed['fedex']['non_document'])}")
    print(f"FedEx per-kg bands: {len(parsed['fedex']['per_kg_bands'])}")


if __name__ == '__main__':
    main()
